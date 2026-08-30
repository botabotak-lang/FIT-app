import re
from pathlib import Path

import openpyxl

from make_template import (
    MATERIAL_SHEETS,
    check_material_sum_formulas,
    collect_worker_names,
    is_formula,
    material_detail_rows,
    scan_residual_values,
    scan_worker_names,
)


ROOT = Path(__file__).resolve().parents[1]
ORIGINAL = ROOT / "docs/reference/FIT様式_作業報告書_材料持出表_原本20260829.xlsx"
TEMPLATE = ROOT / "public/templates/fit_report_template.xlsx"


def main():
    original = openpyxl.load_workbook(ORIGINAL, data_only=False)
    template = openpyxl.load_workbook(TEMPLATE, data_only=False)
    assert original.sheetnames == template.sheetnames
    for name in original.sheetnames:
        o = original[name]
        t = template[name]
        assert len(o.merged_cells.ranges) == len(t.merged_cells.ranges), name
        assert o.print_area == t.print_area, name
        for row in range(11, 15):
            assert o.row_dimensions[row].height == t.row_dimensions[row].height, f"{name} row {row}"
        for col_idx in range(openpyxl.utils.column_index_from_string("B"), openpyxl.utils.column_index_from_string("CC") + 1):
            col = openpyxl.utils.get_column_letter(col_idx)
            assert o.column_dimensions[col].width == t.column_dimensions[col].width, f"{name} col {col}"
    for sheet, cell in [("作業報告書", "CJ138"), ("材料持出表", "M3"), ("材料持出表", "AJ12")]:
        assert original[sheet][cell].value == template[sheet][cell].value, f"{sheet}!{cell}"
    # CJ11 は氏名だけプレースホルダに置き換えてある。氏名以外の構造が原本と同じであること
    def mask_name(value):
        return re.sub(r'="[^"]*"', '="@"', value or "")
    assert mask_name(original["作業報告書"]["CJ11"].value) == mask_name(template["作業報告書"]["CJ11"].value), (
        original["作業報告書"]["CJ11"].value,
        template["作業報告書"]["CJ11"].value,
    )
    assert template["作業報告書"]["CJ11"].value == '=IF(Q11="作業者1",E14-E11,0)', template["作業報告書"]["CJ11"].value

    # 材料持出表系シートの明細行は AJ/AS が必ず数式（原本では値で上書きされた行がある）
    bad = check_material_sum_formulas(template)
    assert not bad, f"合計列が数式でないセル: {bad}"
    for i, name in enumerate(MATERIAL_SHEETS):
        ws = template[name]
        for row in material_detail_rows(i == 0):
            for col in ("AJ", "AS"):
                value = ws[f"{col}{row}"].value
                assert is_formula(value), f"{name}!{col}{row} が数式ではありません: {value!r}"

    # 明細領域・作業報告書ブロック領域に原本の残存値が無いこと
    residual = scan_residual_values(template)
    assert not residual, f"テンプレートに残存値があります: {residual}"

    # 作業者別集計枠に原本の実在社員名が残っていないこと（テンプレートは静的配信される）
    leaked = scan_worker_names(template, collect_worker_names(original))
    assert not leaked, f"テンプレートに原本の氏名が残っています: {leaked[:10]}"

    # テンプレートに作成者名などの個人情報が残っていないこと
    props = template.properties
    for field in ("creator", "lastModifiedBy", "title", "subject", "description", "keywords", "category"):
        value = getattr(props, field, None)
        assert not value, f"template metadata {field}={value!r}"

    print("verify_template.py: OK")
    print("  材料持出表系シートの明細行 AJ/AS: 全て数式")
    print("  明細領域・作業報告書ブロック領域の残存値: 0 件")
    print("  作業者別集計枠の原本氏名: 0 件（プレースホルダ「作業者N」に置換済み）")


if __name__ == "__main__":
    main()
