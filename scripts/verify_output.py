import datetime
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "docs/verify/sample_output.xlsx"
MATERIALS_ONLY = ROOT / "docs/verify/sample_materials.xlsx"
SIX_WORKERS = ROOT / "docs/verify/sample_output_6workers.xlsx"

# render_sample.mjs の既定は架空の作業者4名（作業者A〜作業者D。本番マスタと同じ人数・並び順）
MATERIAL_SHEETS = ["材料持出表", "材料持出表 (2)", "材料持出表 (3)", "材料持出表 (4)"]

# sample_case.json の completionDate（材料持出表 AX1「完成月日」）
EXCEL_EPOCH = datetime.date(1899, 12, 30)
COMPLETION_DATE = datetime.date(2026, 2, 6)
COMPLETION_SERIAL = (COMPLETION_DATE - EXCEL_EPOCH).days

# 合計セル（ラベル AF/AO の右の結合セル）。1ページ目は行25・2ページ目以降は行26
FIRST_TOTAL = ("AJ25", "=SUM(AJ12:AJ24)", "AS25", "=SUM(AS12:AS24)")
REST_TOTAL = ("AJ26", "=SUM(AJ3:AJ25)", "AS26", "=SUM(AS3:AS25)")


def is_formula(value):
    return isinstance(value, str) and value.startswith("=")


def to_date_serial(value):
    """yyyy/m/d 書式のセルは datetime で返るので Excel の日付シリアルに直す"""
    if isinstance(value, datetime.datetime):
        return (value.date() - EXCEL_EPOCH).days
    if isinstance(value, datetime.date):
        return (value - EXCEL_EPOCH).days
    return int(value)


def check_material_totals(wb):
    """材料持出表の仕入合計・売値合計が SUM 数式であること"""
    for i, name in enumerate(MATERIAL_SHEETS):
        ws = wb[name]
        pur_cell, pur_formula, sell_cell, sell_formula = FIRST_TOTAL if i == 0 else REST_TOTAL
        assert ws[pur_cell].value == pur_formula, f"{name}!{pur_cell}={ws[pur_cell].value!r}"
        assert ws[sell_cell].value == sell_formula, f"{name}!{sell_cell}={ws[sell_cell].value!r}"


def to_hours(value):
    """h:mm 書式のセルは timedelta / time で返るので時間数に直す"""
    if isinstance(value, datetime.timedelta):
        return value.total_seconds() / 3600
    if isinstance(value, datetime.time):
        return value.hour + value.minute / 60 + value.second / 3600
    return float(value)


def check_all():
    wb = openpyxl.load_workbook(OUTPUT, data_only=False)
    ws = wb["作業報告書"]
    mat = wb["材料持出表"]

    # 作業報告書：ヘッダー
    assert ws["BU2"].value == "令 和 ８年", ws["BU2"].value
    assert ws["C5"].value == "第一テスト丸"
    assert ws["Y5"].value == "サンプル工事"
    assert ws["BB5"].value == "TEST-100"
    # 製造者の値セルは BU5（BS5:BT8 はラベル）
    assert ws["BU5"].value == "サンプル電機", ws["BU5"].value

    # 作業報告書：1ブロック目（移動 8:00〜9:00）と作業内容の1行目
    assert ws["B11"].value is not None
    assert ws["E11"].value is not None
    assert ws["E14"].value is not None
    assert ws["Q11"].value == "作業者C"
    assert ws["W11"].value == "・機器の取付位置を確認し、取付可否を判断した。"
    # 作業内容は改行で1ブロック1行
    assert ws["W15"].value == "　既存架台の寸法が合わないため、加工が必要と判断した。"
    assert ws["W19"].value == "・必要な部材を確認し、手配を依頼した。"
    # 作業者は社員マスタ順に枠へ入る（1枠目＝作業者A）
    assert ws["CF9"].value == "所要時間　作業者A", ws["CF9"].value
    assert ws["CJ12"].value == '=IF(Q11="作業者A",H14-H11,0)', ws["CJ12"].value
    assert ws["CZ9"].value == "所要時間　作業者C", ws["CZ9"].value
    # 5枠目は社員が居ないので空欄ラベル・0
    assert ws["DT9"].value == "所要時間　", ws["DT9"].value
    assert ws["DX11"].value == "=0", ws["DX11"].value
    # 原本にハードコードされていた氏名が2ページ目以降の見出しに残っていないこと
    for name in ("作業報告書 (2)", "作業報告書 (END)"):
        page = wb[name]
        assert page["CF2"].value == "所要時間　作業者A", f"{name}!CF2={page['CF2'].value}"
        assert page["CF69"].value == "所要時間　作業者A", f"{name}!CF69={page['CF69'].value}"

    # 材料持出表（1ブック統合）：作業者集計は3D参照の数式のまま
    assert mat["D1"].value == "第一テスト丸"
    assert mat["S1"].value == "サンプル工事"
    assert mat["AH1"].value == "TEST-100"
    # 完成月日（AX1）は completionDate。受付日（2026-01-22）ではない
    assert to_date_serial(mat["AX1"].value) == COMPLETION_SERIAL, mat["AX1"].value
    assert mat["A3"].value == "作業者A"
    assert mat["A6"].value == "作業者D"
    assert mat["G3"].value == "=SUM('作業報告書:作業報告書 (END)'!CJ139)", mat["G3"].value
    # 社員は4名なので行7〜9は空欄（0を印字しない）
    for row in (7, 8, 9):
        for col in ("A", "G", "S", "Y", "AK", "M", "AE", "AO"):
            assert mat[f"{col}{row}"].value is None, f"{col}{row}={mat[f'{col}{row}'].value!r}"
    assert mat["A12"].value is not None
    assert mat["E12"].value == "サンプルボルト"
    assert mat["U13"].value == "✓"
    assert mat["AJ12"].value == "=SUM(AF12*AB12)"
    # 単位（Phase C）は AE列
    assert mat["AE12"].value == "本", mat["AE12"].value
    assert mat["AE13"].value == "個", mat["AE13"].value
    # 工賃ヘッダーは設定値の単価
    assert mat["M2"].value == "工　賃(@7,000)", mat["M2"].value
    assert mat["AE2"].value == "工　賃(@8,400)", mat["AE2"].value
    assert mat["AO2"].value == "移動費(×0.8)", mat["AO2"].value
    assert mat["M3"].value == "=(G3*24)*7000", mat["M3"].value
    assert mat["AO3"].value == "=(AK3*24)*5600", mat["AO3"].value

    # 明細行の合計列は全行が数式（原本では952.7等が焼き付いていた）
    for row in range(12, 25):
        assert mat[f"AJ{row}"].value == f"=SUM(AF{row}*AB{row})", f"材料持出表!AJ{row}={mat[f'AJ{row}'].value!r}"
        assert mat[f"AS{row}"].value == f"=SUM(AO{row}*AB{row})", f"材料持出表!AS{row}={mat[f'AS{row}'].value!r}"
    for name in MATERIAL_SHEETS[1:]:
        page = wb[name]
        for row in range(3, 26):
            assert page[f"AJ{row}"].value == f"=SUM(AF{row}*AB{row})", f"{name}!AJ{row}={page[f'AJ{row}'].value!r}"
            assert page[f"AS{row}"].value == f"=SUM(AO{row}*AB{row})", f"{name}!AS{row}={page[f'AS{row}'].value!r}"

    # 仕入合計・売値合計の SUM 数式
    check_material_totals(wb)

    # 14件目以降は2ページ目へ
    assert wb["材料持出表 (2)"]["E3"].value == "サンプル絶縁テープ"
    return ws, mat


def check_materials_only():
    wb = openpyxl.load_workbook(MATERIALS_ONLY, data_only=False)
    assert wb.sheetnames == MATERIAL_SHEETS, wb.sheetnames
    ws = wb["材料持出表"]

    # 作業報告書シートが無いので、3D参照ではなく集計済みの時間値が入る
    for cell in ("G3", "S3", "Y3", "AK3"):
        assert not is_formula(ws[cell].value), f"{cell} は数式のままです: {ws[cell].value!r}"
    # 行3＝作業者A：1/30 の 時間内1h・時間外2h
    assert ws["A3"].value == "作業者A", ws["A3"].value
    assert abs(to_hours(ws["G3"].value) - 1) < 1e-6, ws["G3"].value
    assert abs(to_hours(ws["S3"].value) - 2) < 1e-6, ws["S3"].value
    # 行4＝作業者B：休日3時間
    assert ws["A4"].value == "作業者B", ws["A4"].value
    assert abs(to_hours(ws["Y4"].value) - 3) < 1e-6, ws["Y4"].value
    # 行5＝作業者C：時間内6h・時間外2h・移動1h
    assert ws["A5"].value == "作業者C", ws["A5"].value
    assert abs(to_hours(ws["G5"].value) - 6) < 1e-6, ws["G5"].value
    assert abs(to_hours(ws["S5"].value) - 2) < 1e-6, ws["S5"].value
    assert abs(to_hours(ws["AK5"].value) - 1) < 1e-6, ws["AK5"].value
    # 空き枠（行7〜9）は 0 ではなく空欄
    for row in (7, 8, 9):
        for col in ("A", "G", "S", "Y", "AK", "M", "AE", "AO"):
            assert ws[f"{col}{row}"].value is None, f"{col}{row}={ws[f'{col}{row}'].value!r}"
    # 工賃の数式は残る（値を掛けるだけなので #REF! にならない）
    assert is_formula(ws["M3"].value), ws["M3"].value
    # 船名・科目・型名は値で書かれている（作業報告書への参照ではない）
    assert ws["D1"].value == "第一テスト丸"
    assert ws["S1"].value == "サンプル工事"
    assert to_date_serial(ws["AX1"].value) == COMPLETION_SERIAL, ws["AX1"].value
    # 材料持出表のみの出力でも合計は SUM 数式
    check_material_totals(wb)
    return ws


def check_six_workers():
    """5枠しかないテンプレートに6人目の集計枠（ED/EH）が生成されること"""
    wb = openpyxl.load_workbook(SIX_WORKERS, data_only=False)
    ws = wb["作業報告書"]
    assert ws["ED9"].value == "所要時間　作業者F", ws["ED9"].value
    assert ws["ED136"].value == "合計（作業者F）", ws["ED136"].value
    assert ws["EH11"].value == '=IF(Q11="作業者F",E14-E11,0)', ws["EH11"].value
    assert ws["EH12"].value == '=IF(Q11="作業者F",H14-H11,0)', ws["EH12"].value
    assert ws["ED11"].value == "移動", ws["ED11"].value
    # 合計行と全員合計に6枠目が含まれる
    assert ws["EH138"].value.startswith("=EH11+EH15"), ws["EH138"].value
    assert ws["BZ138"].value == "=CJ138+CT138+DD138+DN138+DX138+EH138", ws["BZ138"].value
    # 作業者名リストは行137から6行
    assert [ws[f"Q{r}"].value for r in range(137, 143)] == [
        "作業者A", "作業者B", "作業者C", "作業者D", "作業者E", "作業者F",
    ], [ws[f"Q{r}"].value for r in range(137, 143)]
    # 2ページ目以降にも6枠目が生成される
    page2 = wb["作業報告書 (2)"]
    assert page2["ED2"].value == "所要時間　作業者F", page2["ED2"].value
    assert page2["EH4"].value == '=IF(Q4="作業者F",E7-E4,0)', page2["EH4"].value
    # 材料持出表は7行しかないので6名は全員入る（行9のみ空欄）
    mat = wb["材料持出表"]
    assert mat["A8"].value == "作業者F", mat["A8"].value
    assert mat["A9"].value is None, mat["A9"].value
    return ws


def main():
    ws, mat = check_all()
    mat_only = check_materials_only()
    six = check_six_workers()
    print("verify_output.py: OK")
    print(f"BU2={ws['BU2'].value} BU5={ws['BU5'].value} W11={ws['W11'].value}")
    print(f"材料持出表(統合) A3={mat['A3'].value} G3={mat['G3'].value} G8={mat['G8'].value!r} G9={mat['G9'].value!r}")
    print(f"材料持出表(単体) G3={mat_only['G3'].value} S3={mat_only['S3'].value} G5={mat_only['G5'].value}")
    print(f"完成月日 AX1={mat['AX1'].value} 合計 AJ25={mat['AJ25'].value} AS25={mat['AS25'].value}")
    print(f"6名レンダー ED9={six['ED9'].value} EH11={six['EH11'].value}")


if __name__ == "__main__":
    main()
