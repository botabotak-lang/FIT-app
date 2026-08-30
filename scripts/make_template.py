"""原本Excelからデータを消してテンプレートを作る。

原本（docs/reference/）はクライアントの実データを含むためリポジトリには含めない。
原本が無い環境ではメッセージを出して終了する（テンプレート自体はコミット済み）。
"""

import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import column_index_from_string, get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "docs/reference/FIT様式_作業報告書_材料持出表_原本20260829.xlsx"
DEST = ROOT / "public/templates/fit_report_template.xlsx"

WORK_SHEETS = [
    "作業報告書",
    "作業報告書 (2)",
    "作業報告書 (3)",
    "作業報告書 (4)",
    "作業報告書 (5)",
    "作業報告書 (END)",
]
MATERIAL_SHEETS = [
    "材料持出表",
    "材料持出表 (2)",
    "材料持出表 (3)",
    "材料持出表 (4)",
]

# 明細行の合計列。原本では大竹氏が数式を値で上書きしたセルがあるため、必ず数式に戻す
MATERIAL_SUM_FORMULAS = {
    "AJ": "=SUM(AF{row}*AB{row})",  # 仕入合計 ＝ 仕入単価 × 数量
    "AS": "=SUM(AO{row}*AB{row})",  # 売値合計 ＝ 売値単価 × 数量
}
# 作業報告書ブロック領域に元から入っている書式用の文字（開始〜終了の区切り）
WORK_REPORT_STATIC_TEXTS = {"～"}

# 作業者別集計枠（印刷範囲外）。原本には実在社員の氏名が見出し・数式に焼き付いている。
# テンプレートは静的配信されるため、氏名をプレースホルダに置き換えてから保存する。
# アプリ側（lib/reportWorkbook.ts writeWorkerFormulas）が出力時に社員マスタの氏名で上書きする。
WORKER_LABEL_COLS = ["CF", "CP", "CZ", "DJ", "DT"]
WORKER_VALUE_COLS = ["CJ", "CT", "DD", "DN", "DX"]
WORKER_TOTAL_LABEL_ROW = 136
WORKER_NAME_LIST_COL = "Q"
WORKER_NAME_LIST_ROWS = range(137, 142)


def worker_placeholder(slot):
    """1始まりの枠番号でプレースホルダ氏名を作る"""
    return f"作業者{slot + 1}"


def material_detail_rows(first_page):
    return range(12, 25) if first_page else range(3, 26)


def work_report_block_starts(first_page):
    if first_page:
        return list(range(11, 67, 4)) + list(range(70, 134, 4))
    return list(range(4, 68, 4)) + list(range(71, 135, 4))


def worker_header_rows(first_page):
    """作業者別集計の見出し行（ページ上部・ページ中央）。シートで位置が違う"""
    return [9, 68] if first_page else [2, 69]


def clear_if_not_formula(ws, cell):
    if isinstance(ws[cell], MergedCell):
        return
    value = ws[cell].value
    if isinstance(value, str) and value.startswith("="):
        return
    ws[cell].value = None


def clear_work_report(ws, first_page):
    if first_page:
        for cell in ("C5", "Y5", "BB5", "BT5", "BU5", "BU2"):
            clear_if_not_formula(ws, cell)
    starts = work_report_block_starts(first_page)

    for row in starts:
        for col in ("B", "E", "H", "K", "N", "Q", "T", "W"):
            clear_if_not_formula(ws, f"{col}{row}")
        for col in ("E", "H", "K", "N"):
            clear_if_not_formula(ws, f"{col}{row + 3}")


def clear_materials(ws, first_page):
    if first_page:
        for cell in ("D1", "S1", "AH1", "AX1"):
            clear_if_not_formula(ws, cell)
        for row in range(3, 10):
            clear_if_not_formula(ws, f"A{row}")
            clear_if_not_formula(ws, f"AU{row}")
        carrier_cell = "Z25"
    else:
        carrier_cell = "Z26"
    detail_rows = material_detail_rows(first_page)

    for row in detail_rows:
        for col in ("A", "E", "M", "U", "X", "AB", "AE", "AF", "AO", "AX"):
            clear_if_not_formula(ws, f"{col}{row}")
        # 原本で数式が値に置き換わっている行があるため、合計列は無条件に数式へ戻す
        for col, formula in MATERIAL_SUM_FORMULAS.items():
            ws[f"{col}{row}"] = formula.format(row=row)
    clear_if_not_formula(ws, carrier_cell)


def collect_worker_names(wb):
    """原本の作業者別集計枠に焼き付いている実在社員の氏名を集める"""
    names = set()
    ws = wb["作業報告書"]
    for col in WORKER_LABEL_COLS:
        value = ws[f"{col}9"].value
        if isinstance(value, str) and "　" in value:
            name = value.split("　", 1)[1].strip()
            if name:
                names.add(name)
    for row in WORKER_NAME_LIST_ROWS:
        value = ws[f"{WORKER_NAME_LIST_COL}{row}"].value
        if isinstance(value, str) and value.strip():
            names.add(value.strip())
    for row in range(3, 10):
        value = wb["材料持出表"][f"A{row}"].value
        if isinstance(value, str) and value.strip():
            names.add(value.strip())
    return names


def neutralize_worker_names(ws, first_page):
    """作業者別集計枠の氏名を「作業者N」に置き換える（見出し・合計ラベル・数式・氏名リスト）"""
    starts = work_report_block_starts(first_page)
    for slot, (label_col, value_col) in enumerate(zip(WORKER_LABEL_COLS, WORKER_VALUE_COLS)):
        placeholder = worker_placeholder(slot)
        for row in worker_header_rows(first_page):
            ws[f"{label_col}{row}"] = f"所要時間　{placeholder}"
        ws[f"{label_col}{WORKER_TOTAL_LABEL_ROW}"] = f"合計（{placeholder}）"
        for start in starts:
            for offset in range(4):
                cell = ws[f"{value_col}{start + offset}"]
                if is_formula(cell.value):
                    # =IF(Q11="大竹",E14-E11,0) の氏名部分だけを差し替える
                    cell.value = re.sub(r'="[^"]*"', f'="{placeholder}"', cell.value, count=1)
    for slot, row in enumerate(WORKER_NAME_LIST_ROWS):
        ws[f"{WORKER_NAME_LIST_COL}{row}"] = worker_placeholder(slot)


def scan_worker_names(wb, names):
    """氏名がブックのどこかに残っていないか全シート走査する"""
    found = []
    if not names:
        return found
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                for name in names:
                    if name in value:
                        found.append((ws.title, cell.coordinate, value))
                        break
    return found


def is_formula(value):
    return isinstance(value, str) and value.startswith("=")


def _scan_range(ws, rows, first_col, last_col, allowed_texts=()):
    """数式でも空でもないセル（＝原本の残存値）を列挙する"""
    found = []
    for row in rows:
        for idx in range(column_index_from_string(first_col), column_index_from_string(last_col) + 1):
            cell = ws.cell(row=row, column=idx)
            if isinstance(cell, MergedCell):
                continue
            value = cell.value
            if value is None or is_formula(value) or value in allowed_texts:
                continue
            found.append((ws.title, f"{get_column_letter(idx)}{row}", value))
    return found


def scan_residual_values(wb):
    """材料持出表の明細領域・作業報告書のブロック領域に残っている値を列挙する。

    どちらもアプリが毎回書き込む領域なので、テンプレートには数式か空欄しか残っていてはいけない。
    """
    found = []
    for i, name in enumerate(MATERIAL_SHEETS):
        if name not in wb.sheetnames:
            continue
        # 印刷範囲 A1:BB25 の明細行を丸ごと見る（AJ17 のような列外の焼き付きを拾うため）
        found += _scan_range(wb[name], material_detail_rows(i == 0), "A", "BB")
    for i, name in enumerate(WORK_SHEETS):
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        rows = [r + offset for r in work_report_block_starts(i == 0) for offset in range(4)]
        # 印刷範囲 B2:CC133。CF列以降は作業者別集計（ラベル文字が入る）ので対象外
        found += _scan_range(ws, rows, "B", "CC", WORK_REPORT_STATIC_TEXTS)
    return found


def check_material_sum_formulas(wb):
    """明細行の AJ/AS が全て所定の数式であること"""
    bad = []
    for i, name in enumerate(MATERIAL_SHEETS):
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        for row in material_detail_rows(i == 0):
            for col, formula in MATERIAL_SUM_FORMULAS.items():
                expected = formula.format(row=row)
                actual = ws[f"{col}{row}"].value
                if actual != expected:
                    bad.append((name, f"{col}{row}", actual, expected))
    return bad


def clear_metadata(wb):
    """作成者名などの個人情報をテンプレートに残さない"""
    props = wb.properties
    props.creator = ""
    props.lastModifiedBy = ""
    props.title = ""
    props.subject = ""
    props.description = ""
    props.keywords = ""
    props.category = ""
    props.identifier = None
    props.language = None
    props.revision = None
    props.contentStatus = None


def main():
    if not SOURCE.exists():
        print(f"原本が見つかりません: {SOURCE}")
        print("原本は実データを含むためリポジトリに含めていません。")
        print("テンプレート（public/templates/fit_report_template.xlsx）はコミット済みのものを使用してください。")
        return 0
    wb = openpyxl.load_workbook(SOURCE)
    worker_names = collect_worker_names(wb)
    for i, name in enumerate(WORK_SHEETS):
        clear_work_report(wb[name], i == 0)
        neutralize_worker_names(wb[name], i == 0)
    for i, name in enumerate(MATERIAL_SHEETS):
        clear_materials(wb[name], i == 0)
    clear_metadata(wb)

    bad = check_material_sum_formulas(wb)
    if bad:
        print("材料持出表の合計列が数式になっていません:")
        for sheet, cell, actual, expected in bad:
            print(f"  {sheet}!{cell} = {actual!r}（期待: {expected}）")
        return 1
    residual = scan_residual_values(wb)
    if residual:
        print(f"テンプレートに原本の残存値が {len(residual)} 件あります:")
        for sheet, cell, value in residual:
            print(f"  {sheet}!{cell} = {value!r}")
        return 1
    leaked = scan_worker_names(wb, worker_names)
    if leaked:
        print(f"テンプレートに原本の氏名が {len(leaked)} 件残っています:")
        for sheet, cell, value in leaked[:20]:
            print(f"  {sheet}!{cell} = {value!r}")
        return 1

    DEST.parent.mkdir(parents=True, exist_ok=True)
    wb.save(DEST)
    print(f"created {DEST}")
    print("残存値スキャン: 0 件（明細領域・作業報告書ブロック領域）")
    print(f"氏名スキャン: 0 件（原本の氏名 {len(worker_names)} 名を全シート走査）")
    return 0


if __name__ == "__main__":
    sys.exit(main())
